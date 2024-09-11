import pandas as pd
import openpyxl

def beneficio_pleno_escribir(datos_beneficio_pleno, ruta_destino, nombre_archivo, ubicacion_formato):
    start_time = pd.Timestamp.now()

    # Cargar la plantilla de Excel usando openpyxl
    print(f"Intentando cargar el archivo de Excel: {ubicacion_formato}")
    try:
        libro_excel = openpyxl.load_workbook(ubicacion_formato)
        print("Archivo de Excel cargado con éxito.")
    except Exception as e:
        print(f"Error al cargar el archivo de Excel: {e}")
        return

    # Seleccionar la hoja de cálculo donde se van a escribir los datos
    hoja_calculo = libro_excel['BENEF PLENO VIT.(4)']

    # Leer los datos del archivo CSV
    try:
        print("Leyendo el archivo CSV...")
        datos_beneficio_csv = pd.read_csv(datos_beneficio_pleno, sep=';', encoding='latin1')
    except Exception as e:
        print(f"Error al leer el archivo CSV: {e}")
        return

    # Escribir las primeras 6 columnas en el archivo Excel
    print("Escribiendo los datos (primeras 6 columnas)...")
    for idx, row in datos_beneficio_csv.iterrows():
        for col_idx, value in enumerate(row[:6], start=1):  # Las primeras 6 columnas
            hoja_calculo.cell(row=idx+6, column=col_idx, value=value)

    # Escribir las columnas 8, 9 y 10 en el archivo Excel
    print("Escribiendo las columnas 8, 9 y 10...")
    for idx, row in datos_beneficio_csv.iterrows():
        for col_idx, value in enumerate(row[7:10], start=8):  # Columnas 8, 9 y 10
            hoja_calculo.cell(row=idx+6, column=col_idx, value=value)

    # Escribir la columna 11 en el archivo Excel
    print("Escribiendo la columna 11...")
    for idx, value in enumerate(datos_beneficio_csv.iloc[:, 10], start=6):
        hoja_calculo.cell(row=idx+1, column=11, value=value)

    # Aplicar fórmulas
    print("Aplicando fórmulas...")
    num_filas = len(datos_beneficio_csv) + 6
    for fila in range(6, num_filas + 1):
        hoja_calculo.cell(row=fila, column=7).value = f'=+INT(($B$3-F{fila})/365.25+0.5)'
        hoja_calculo.cell(row=fila, column=11).value = f'=IF(I{fila}=1,J{fila},IF(I{fila}=14,J{fila}-H{fila},0))'
        hoja_calculo.cell(row=fila, column=13).value = f'=IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(G{fila},HOMBRES,11,0),VLOOKUP(G{fila},MUJERES,11,0)),IF(D{fila}="M",VLOOKUP(G{fila},Homb_inv,11,0),VLOOKUP(G{fila},Mujer_inv,11,0)))'
        hoja_calculo.cell(row=fila, column=14).value = f'=IF($D$1=1,IF(D{fila}="F",VLOOKUP(#REF!,HOMBRES,11,0),VLOOKUP(#REF!,MUJERES,11,0)),0)'    
        hoja_calculo.cell(row=fila, column=15).value = f'=IF($D$1=1,IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(G{fila},axy_,#REF!-13,0),HLOOKUP(G{fila},axy_,#REF!-13,0)),IF(D{fila}="M",VLOOKUP(G{fila},axy__hi__mv,#REF!,0),HLOOKUP(G{fila},ayx__mi_hv,#REF!,0))),0)'
        hoja_calculo.cell(row=fila, column=16).value = f'=MAX(0,N{fila}-O{fila})'
        hoja_calculo.cell(row=fila, column=17).value = f'=IF($D$1=1,IF(H{fila}<5*$B$2,5*$B$2,IF(H{fila}>10*$B$2,10*$B$2,H{fila})),0)'
        hoja_calculo.cell(row=fila, column=18).value = f'=IF($D$1=1,IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(#REF!,HOMBRES,14,0),VLOOKUP(#REF!,MUJERES,14,0)),IF(D{fila}="M",VLOOKUP(#REF!,Homb_inv,14,0),VLOOKUP(#REF!,Mujer_inv,14,0))),0)'
        hoja_calculo.cell(row=fila, column=19).value = f'=12*(M{fila}*TABLA!$H$8+TABLA!$H$9)+2*(M{fila}*TABLA!$H$10+TABLA!$H$11)'
        hoja_calculo.cell(row=fila, column=20).value = f'=12*(M{fila}*TABLA!$H$8+TABLA!$H$9)+M{fila}'
        hoja_calculo.cell(row=fila, column=21).value = f'=12*(M{fila}*TABLA!$H$8+TABLA!$H$9)'
        hoja_calculo.cell(row=fila, column=22).value = f'=2*(M{fila}*TABLA!$H$10+TABLA!$H$11)-M{fila}'
        hoja_calculo.cell(row=fila, column=23).value = f'=(12*P{fila}*TABLA!$H$8)+2*P{fila}*TABLA!$H$10'
        hoja_calculo.cell(row=fila, column=24).value = f'=(12*P{fila}*TABLA!$H$8)+P{fila}'
        hoja_calculo.cell(row=fila, column=25).value = f'=(12*P{fila}*TABLA!$H$8)'
        hoja_calculo.cell(row=fila, column=26).value = f'=(2*P{fila}*TABLA!$H$10)-P{fila}'
        hoja_calculo.cell(row=fila, column=27).value = f'=MAX(0,H{fila})*IF(I{fila}=14,S{fila},T{fila})'
        hoja_calculo.cell(row=fila, column=28).value = f'=MAX(0,H{fila})*IF(I{fila}=14,W{fila},X{fila})'
        hoja_calculo.cell(row=fila, column=29).value = f'=MAX(0,H{fila})*IF(I{fila}=14,W{fila},X{fila})'
        hoja_calculo.cell(row=fila, column=30).value = f'=K{fila}*Z{fila}'
        hoja_calculo.cell(row=fila, column=31).value = f'=U{fila}*L{fila}'
        hoja_calculo.cell(row=fila, column=32).value = f'=Y{fila}*L{fila}'
        hoja_calculo.cell(row=fila, column=33).value = f'=(Q{fila}*R{fila})'
        hoja_calculo.cell(row=fila, column=34).value = f'=SUM(AA{fila}:AG{fila})'
        hoja_calculo.cell(row=fila, column=36).value = f'=AA{fila}+AB{fila}'
        hoja_calculo.cell(row=fila, column=37).value = f'=AC{fila}+AD{fila}'
        hoja_calculo.cell(row=fila, column=38).value = f'=AG{fila}'
        hoja_calculo.cell(row=fila, column=39).value = f'=SUM(AJ{fila}:AL{fila})'
        hoja_calculo.cell(row=fila, column=40).value = f'=AE{fila}+AF{fila}'
        
    # Guardar los cambios en el archivo Excel
    print("Guardando el archivo con las fórmulas aplicadas...")
    try:
        ruta_archivo_salida = ruta_destino + nombre_archivo + '.xlsx'
        libro_excel.save(ruta_archivo_salida)
        print(f"Libro guardado con éxito en {ruta_archivo_salida}.")
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")

    end_time = pd.Timestamp.now()
    tiempo_ejecucion = (end_time - start_time).total_seconds() / 60
    print(f"Tiempo de ejecución: {tiempo_ejecucion} minutos")     