import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def plenas_escribir(datos_plenas, ruta_destino, nombre_archivo, ubicacion_formato):
    start_time = pd.Timestamp.now()

    # Cargar la plantilla de Excel con openpyxl
    print("Cargando la plantilla de Excel")
    libro_excel = openpyxl.load_workbook(ubicacion_formato)

    # Selecciona la hoja de cálculo a modificar
    hoja_calculo = libro_excel['CALCULO PLENAS (1)']

    # Leer los datos del archivo CSV
    print("Lectura de los datos del archivo CSV")
    datos_plenas_csv = pd.read_csv(datos_plenas, sep=';', encoding='latin1')

    # Escribir las primeras 6 columnas directamente en la hoja usando openpyxl
    print("Escribiendo las primeras 6 columnas")
    for idx, row in enumerate(dataframe_to_rows(datos_plenas_csv.iloc[:, :6], index=False, header=False), start=7):
        for col_idx, value in enumerate(row, start=1):
            hoja_calculo.cell(row=idx, column=col_idx, value=value)

    # Escribir la columna 7 en el excel columna 8
    print("Escribiendo la columna 7")
    for idx, value in enumerate(datos_plenas_csv.iloc[:, 6], start=7):
        hoja_calculo.cell(row=idx, column=8, value=value)

    # Escribir las columnas 8, 9 y 10 del CSV en las columnas 10, 11 y 12 del Excel, comenzando desde la fila 7
    print("Escribiendo las columnas 8, 9 y 10")
    for idx, row in enumerate(dataframe_to_rows(datos_plenas_csv.iloc[:, 7:10], index=False, header=False), start=7):
        for col_idx, value in enumerate(row, start=10):
            hoja_calculo.cell(row=idx, column=col_idx, value=value)


    # Escribir la columna 11 del csv en la columna 14 del Excel
    print("Escribiendo la columna 11")
    for idx, value in enumerate(datos_plenas_csv.iloc[:, 10], start=7):
        hoja_calculo.cell(row=idx, column=14, value=value)

    # Escribir las fórmulas en las filas
    print("Escribiendo fórmulas en las filas")
    num_filas = len(datos_plenas_csv) + 6
    for fila in range(7, num_filas + 1):  # Cambiamos el inicio del rango a 7
        hoja_calculo.cell(row=fila, column=7).value = f'=INT(($B$3-F{fila})/365.25+0.5)'
        hoja_calculo.cell(row=fila, column=9).value = f'=INT(($B$3-H{fila})/365.25+0.5)'
        hoja_calculo.cell(row=fila, column=12).value = f'=J{fila}'
        hoja_calculo.cell(row=fila, column=13).value = f'=IF(K{fila}=1,L{fila},IF(K{fila}=14,L{fila}-J{fila},0))'
        hoja_calculo.cell(row=fila, column=15).value = f'=IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(G{fila},HOMBRES,11,0),VLOOKUP(G{fila},MUJERES,11,0)),IF(D{fila}="M",VLOOKUP(G{fila},Homb_inv,11,0),VLOOKUP(G{fila},Mujer_inv,11,0)))'
        hoja_calculo.cell(row=fila, column=16).value = f'=IF($D$1=1,IF(D{fila}="F",VLOOKUP(I{fila},HOMBRES,11,0),VLOOKUP(I{fila},MUJERES,11,0)),0)'
        hoja_calculo.cell(row=fila, column=17).value = f'=IF($D$1=1,IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(G{fila},axy_,I{fila}-13,0),HLOOKUP(G{fila},axy_,I{fila}-13,0)),IF(D{fila}="M",VLOOKUP(G{fila},axy__hi__mv,I{fila}-13,0),HLOOKUP(G{fila},ayx__mi__hv,I{fila}-13,0))),0)'
        hoja_calculo.cell(row=fila, column=18).value = f'=MAX(0,P{fila}-Q{fila})'
        hoja_calculo.cell(row=fila, column=19).value = f'=IF(J{fila}<5*$B$2,5*$B$2,IF(J{fila}>10*$B$2,10*$B$2,J{fila}))'
        hoja_calculo.cell(row=fila, column=20).value = f'=IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(G{fila},HOMBRES,14,0),VLOOKUP(G{fila},MUJERES,14,0)),IF(D{fila}="M",VLOOKUP(G{fila},Homb_inv,14,0),VLOOKUP(G{fila},Mujer_inv,14,0)))'
        hoja_calculo.cell(row=fila, column=21).value = f'=12*(O{fila}*TABLA!$H$8+TABLA!$H$9)+2*(O{fila}*TABLA!$H$10+TABLA!$H$11)'
        hoja_calculo.cell(row=fila, column=22).value = f'=12*(O{fila}*TABLA!$H$8+TABLA!$H$9)+1*(O{fila}*TABLA!$H$10+TABLA!$H$11)'
        hoja_calculo.cell(row=fila, column=23).value = f'=12*(O{fila}*TABLA!$H$8+TABLA!$H$9)'
        hoja_calculo.cell(row=fila, column=24).value = f'=1*(O{fila}*TABLA!$H$10+TABLA!$H$11)'
        hoja_calculo.cell(row=fila, column=25).value = f'=(12*R{fila}*TABLA!$H$8)+2*R{fila}*TABLA!$H$10'
        hoja_calculo.cell(row=fila, column=26).value = f'=(12*R{fila}*TABLA!$H$8)+1*R{fila}*TABLA!$H$10'
        hoja_calculo.cell(row=fila, column=27).value = f'=(12*R{fila}*TABLA!$H$8)'
        hoja_calculo.cell(row=fila, column=28).value = f'=(1*R{fila}*TABLA!$H$10)'
        hoja_calculo.cell(row=fila, column=29).value = f'=MAX(0,J{fila})*IF(K{fila}=14,U{fila},V{fila})'
        hoja_calculo.cell(row=fila, column=30).value = f'=X{fila}*M{fila}'
        hoja_calculo.cell(row=fila, column=31).value = f'=MAX(0,J{fila})*IF(K{fila}=14,Y{fila},Z{fila})'
        hoja_calculo.cell(row=fila, column=32).value = f'=M{fila}*AB{fila}'
        hoja_calculo.cell(row=fila, column=33).value = f'=W{fila}*N{fila}'
        hoja_calculo.cell(row=fila, column=34).value = f'=AA{fila}*N{fila}'
        hoja_calculo.cell(row=fila, column=35).value = f'=(S{fila}*T{fila})'
        hoja_calculo.cell(row=fila, column=36).value = f'=SUM(AC{fila}:AI{fila})'
        hoja_calculo.cell(row=fila, column=38).value = f'=AC{fila}+AD{fila}'
        hoja_calculo.cell(row=fila, column=39).value = f'=AE{fila}+AF{fila}'
        hoja_calculo.cell(row=fila, column=40).value = f'=AI{fila}'
        hoja_calculo.cell(row=fila, column=41).value = f'=SUM(AL{fila}:AN{fila})'
        hoja_calculo.cell(row=fila, column=42).value = f'=AG{fila}+AH{fila}'

    # Guardar el archivo Excel con los cambios
    print("Guardando el archivo Excel con los cambios")
    libro_excel.save(ruta_destino + nombre_archivo + '.xlsx')

    end_time = pd.Timestamp.now()
    tiempo_ejecucion = (end_time - start_time).total_seconds() / 60
    print(f"Tiempo de ejecución: {tiempo_ejecucion} minutos")
    print("Libro guardado con éxito")
