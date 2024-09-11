import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import numbers

def compartidas_escribir(datos_compartidas, ruta_destino, nombre_archivo, ubicacion_formato):
    start_time = pd.Timestamp.now()

    # Verificar si el archivo de plantilla de Excel existe
    if not os.path.exists(ubicacion_formato):
        print(f"Error: El archivo de plantilla {ubicacion_formato} no existe.")
        return

    # Verificar si el archivo tiene la extensión correcta
    if not ubicacion_formato.endswith('.xlsx'):
        print(f"Error: El archivo {ubicacion_formato} no es un archivo .xlsx válido.")
        return

    try:
        # Intentar cargar la plantilla de Excel
        print(f"Intentando cargar el archivo de Excel: {ubicacion_formato}")
        libro_excel = openpyxl.load_workbook(ubicacion_formato)
        print("Archivo de Excel cargado con éxito.")
    except Exception as e:
        print(f"Error al cargar la plantilla de Excel: {e}")
        return

    try:
        # Crear la ruta para el archivo de salida
        ruta_archivo_salida = os.path.join(ruta_destino, nombre_archivo + '.xlsx')

        # Verificar si el directorio de destino existe, si no, crearlo
        if not os.path.exists(ruta_destino):
            os.makedirs(ruta_destino)
            print(f"Directorio {ruta_destino} creado.")

        # Leer los datos del archivo CSV
        print("Lectura del archivo CSV")
        datos_compartidas_csv = pd.read_csv(datos_compartidas, sep=';', encoding='latin1')

        # Seleccionar la hoja de cálculo
        hoja_calculo = libro_excel['CALCULO COMPARTIDAS(2)']

        # Escribir las primeras 5 columnas del DataFrame en el Excel, comenzando desde la columna 1 y la fila 6
        print("Escribiendo los datos (primeras 5 columnas)")
        for idx, row in datos_compartidas_csv.iterrows():
            for col_idx, value in enumerate(row[:5], start=1):  # Comenzar desde la columna 1 en Excel
                hoja_calculo.cell(row=idx+6, column=col_idx, value=value)  # Comenzar a escribir en la fila 6


        # Mapeo de las columnas del DataFrame a las columnas correspondientes del Excel
        column_mappings = {
            5: 6,    # Columna 6 del DataFrame -> Columna 6 del Excel
            6: 8,    # Columna 7 del DataFrame -> Columna 8 del Excel
            7: 10,   # Columna 8 del DataFrame -> Columna 10 del Excel
            8: 11,   # Columna 9 del DataFrame -> Columna 11 del Excel
            9: 14,   # Columna 10 del DataFrame -> Columna 14 del Excel
            10: 19   # Columna 11 del DataFrame -> Columna 19 del Excel
            }

        # Iterar sobre las columnas del DataFrame y escribirlas en las columnas correspondientes del Excel
        for df_col, excel_col in column_mappings.items():
            print(f"Escribiendo la columna {df_col + 1} del DataFrame en la columna {excel_col} del Excel")
            for idx, value in enumerate(datos_compartidas_csv.iloc[:, df_col], start=6):
                hoja_calculo.cell(row=idx, column=excel_col, value=value)

        
        # Aplicar fórmulas
        num_filas = len(datos_compartidas_csv) + 5
        for fila in range(6, num_filas + 1):
            hoja_calculo.cell(row=fila, column=7).value = f'=INT(($B$3-F{fila})/365.25+0.5)'
            hoja_calculo.cell(row=fila, column=7).number_format = numbers.FORMAT_NUMBER
            hoja_calculo.cell(row=fila, column=9).value = f'=IF(H{fila}>0,(INT(($B$3-H{fila})/365.25+0.5)),IF(D{fila}="M",G{fila}-5,G{fila}+5))'
            hoja_calculo.cell(row=fila, column=12).value = f'=IF(K{fila}=14,J{fila},0)'
            hoja_calculo.cell(row=fila, column=13).value = f'=IF(K{fila}=1,L{fila},IF(K{fila}=14,J{fila}-L{fila},0))'
            hoja_calculo.cell(row=fila, column=15).value = f'=IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(G{fila},HOMBRES,11,0),VLOOKUP(G{fila},MUJERES,11,0)),IF(D{fila}="M",VLOOKUP(G{fila},Homb_inv,11,0),VLOOKUP(G{fila},Mujer_inv,11,0)))'
            hoja_calculo.cell(row=fila, column=16).value = f'=IF(D{fila}="F",VLOOKUP(I{fila},HOMBRES,11,0),VLOOKUP(I{fila},MUJERES,11,0))'
            hoja_calculo.cell(row=fila, column=17).value = f'=IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(G{fila},axy_,I{fila}-13,0),HLOOKUP(G{fila},axy_,I{fila}-13,0)),IF(D{fila}="M",VLOOKUP(G{fila},axy__hi__mv,I{fila}-13,0),HLOOKUP(G{fila},ayx__mi__hv,I{fila}-13,0)))'
            hoja_calculo.cell(row=fila, column=18).value = f'=P{fila}-Q{fila}'
            hoja_calculo.cell(row=fila, column=20).value = f'=IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(G{fila},HOMBRES,14,0),VLOOKUP(G{fila},MUJERES,14,0)),IF(D{fila}="M",VLOOKUP(G{fila},Homb_inv,14,0),VLOOKUP(G{fila},Mujer_inv,14,0)))'
            hoja_calculo.cell(row=fila, column=21).value = f'=12*(O{fila}*TABLA!$H$8+TABLA!$H$9)+2*(O{fila}*TABLA!$H$10+TABLA!$H$11)'
            hoja_calculo.cell(row=fila, column=22).value = f'=12*(O{fila}*TABLA!$H$8+TABLA!$H$9)+1*(O{fila}*TABLA!$H$10+TABLA!$H$11)'
            hoja_calculo.cell(row=fila, column=23).value = f'=12*(O{fila}*TABLA!$H$8+TABLA!$H$9)'
            hoja_calculo.cell(row=fila, column=24).value = f'=1*(O{fila}*TABLA!$H$10+TABLA!$H$11)'
            hoja_calculo.cell(row=fila, column=25).value = f'=(12*R{fila}*TABLA!$H$8)+2*R{fila}*TABLA!$H$10'
            hoja_calculo.cell(row=fila, column=26).value = f'=(12*R{fila}*TABLA!$H$8)+1*R{fila}*TABLA!$H$10'
            hoja_calculo.cell(row=fila, column=27).value = f'=(12*R{fila}*TABLA!$H$8)'
            hoja_calculo.cell(row=fila, column=28).value = f'=(1*R{fila}*TABLA!$H$10)'
            hoja_calculo.cell(row=fila, column=29).value = f'=MAX(0,J{fila})*IF(K{fila}=14,U{fila},V{fila})'
            hoja_calculo.cell(row=fila, column=30).value = f'=X{fila}*M{fila}'
            hoja_calculo.cell(row=fila, column=31).value = f'=MAX(0,J{fila})*IF(K{fila}=14,Y{fila},Z{fila})'
            hoja_calculo.cell(row=fila, column=32).value = f'=M{fila}*AB{fila}'
            hoja_calculo.cell(row=fila, column=33).value = f'=W{fila}*N{fila}'
            hoja_calculo.cell(row=fila, column=34).value = f'=AA{fila}*N{fila}'
            hoja_calculo.cell(row=fila, column=35).value = f'=(S{fila}*T{fila})'
            hoja_calculo.cell(row=fila, column=36).value = f'=AC{fila}+AE{fila}+AI{fila}+AG{fila}+AD{fila}+AF{fila}+AH{fila}'
            hoja_calculo.cell(row=fila, column=38).value = f'=AC{fila}+AD{fila}'
            hoja_calculo.cell(row=fila, column=39).value = f'=AE{fila}+AF{fila}'
            hoja_calculo.cell(row=fila, column=40).value = f'=AI{fila}'
            hoja_calculo.cell(row=fila, column=41).value = f'=SUM(AL{fila}:AN{fila})'
            hoja_calculo.cell(row=fila, column=42).value = f'=AG{fila}+AH{fila}'
 
        # Guardar el archivo Excel
        print(f"Guardando el archivo con las fórmulas aplicadas en {ruta_archivo_salida}...")
        libro_excel.save(ruta_archivo_salida)
        print("Archivo guardado con éxito.")

    except Exception as e:
       print(f"Error durante la escritura o guardado del archivo Excel: {e}")
       return

    end_time = pd.Timestamp.now()
    tiempo_ejecucion = (end_time - start_time).total_seconds() / 60
    print(f"Tiempo de ejecución: {tiempo_ejecucion} minutos")
    
