import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def jub_escribir(datos_jub, ruta_destino, nombre_archivo, ubicacion_formato):
    start_time = pd.Timestamp.now()

    # Verificar la ruta completa del archivo de plantilla
    print(f"Ruta del archivo de plantilla: {ubicacion_formato}")
    
    # Intentar cargar la plantilla de Excel
    try:
        libro_excel = openpyxl.load_workbook(ubicacion_formato)
        print("Archivo de plantilla cargado correctamente.")
    except Exception as e:
        print(f"Error al cargar el archivo de plantilla: {e}")
        return  # Termina la función si no se puede cargar el archivo

    # Leer los datos del archivo CSV
    try:
        print("Lectura del archivo CSV")
        datos_jub_csv = pd.read_csv(datos_jub, sep=';', encoding='latin1')
        print("Archivo CSV cargado correctamente.")
    except Exception as e:
        print(f"Error al leer el archivo CSV: {e}")
        return

    # Ahora escribimos en el archivo Excel directamente sin usar pd.ExcelWriter
    try:
        hoja_calculo = libro_excel['Jub en expectativa(3)']

        # Escribir las primeras 6 columnas
        print("Escribiendo los datos (primeras 6 columnas)")
        for idx, row in enumerate(dataframe_to_rows(datos_jub_csv.iloc[:, :6], index=False, header=False), start=5):
            for col_idx, value in enumerate(row, start=1):
                hoja_calculo.cell(row=idx, column=col_idx).value = value

        # Definir las columnas a escribir (origen -> destino en Excel)
        column_mappings = {
            7: 20,  # Columna 8 del CSV a la columna 20 del Excel
            8: 21,  # Columna 9 del CSV a la columna 21 del Excel
            9: 22,  # Columna 10 del CSV a la columna 22 del Excel
            10: 23, # Columna 11 del CSV a la columna 23 del Excel
            11: 25, # Columna 12 del CSV a la columna 25 del Excel
            12: 14  # Columna 13 del CSV a la columna 14 del Excel
        }

        # Iterar sobre los pares columna origen (CSV) -> columna destino (Excel)
        for csv_col, excel_col in column_mappings.items():
            print(f"Escribiendo la columna {csv_col + 1} en la columna {excel_col} del Excel")  # csv_col + 1 para mostrar 1-based index
            for idx, value in enumerate(datos_jub_csv.iloc[:, csv_col], start=5):  # Comenzar desde la fila 5 en Excel
                hoja_calculo.cell(row=idx, column=excel_col).value = value  # Escribir el valor en la columna de destino
    
        
        
        # Escribir las fórmulas
        num_filas = len(datos_jub_csv) + 4
        print("Escribiendo las fórmulas")
        for fila in range(5, num_filas + 1):
            hoja_calculo.cell(row=fila, column=7).value = f'=INT(($B$3-F{fila})/365.25+0.5)'
            hoja_calculo.cell(row=fila, column=9).value = f'=INT(IF(H{fila}>0,(($B$3-H{fila})/365.25+0.5),IF(D{fila}="M",G{fila}-5,G{fila}+5)))'
            hoja_calculo.cell(row=fila, column=10).value = f'=F{fila}+P{fila}*365.25'
            hoja_calculo.cell(row=fila, column=11).value = f'=MAX(0,ROUNDUP((J{fila}-$B$3)/30.4375,0))'
            hoja_calculo.cell(row=fila, column=12).value = f'=U{fila}'
            hoja_calculo.cell(row=fila, column=13).value = f'=L{fila}*IF(L{fila}/$B$2>4,0.17,0.16)'
            hoja_calculo.cell(row=fila, column=15).value = f'=N{fila}+(J{fila}-$B$3)/7'
            hoja_calculo.cell(row=fila, column=16).value = f'=MAX(G{fila},IF(D{fila}=\"M\",IF(DATE(2015,1,1)>365.25*60+F{fila},60,62),IF(DATE(2015,1,1)>365.25*55+F{fila},55,57)),ROUNDUP(G{fila}+(1300-N{fila})*7/365.25+0.5,0))'
            hoja_calculo.cell(row=fila, column=17).value = f'=P{fila}-G{fila}'
            hoja_calculo.cell(row=fila, column=18).value = f'=P{fila}-(G{fila}-I{fila})'
            hoja_calculo.cell(row=fila, column=19).value = f'=IF(((0.655-(L{fila}/$B$2)*0.005)*L{fila})<$B$2,$B$2,(0.655-(L{fila}/$B$2)*0.005)*L{fila})'
            hoja_calculo.cell(row=fila, column=24).value = f'=IF(V{fila}=1,W{fila},IF(V{fila}=14,W{fila}-U{fila},0))'
            hoja_calculo.cell(row=fila, column=26).value = f'=IF(E{fila}=\"NO\",IF(D{fila}=\"M\",VLOOKUP(G{fila},HOMBRES,11,0),VLOOKUP(G{fila},MUJERES,11,0)),IF(C{fila}=\"M\",VLOOKUP(G{fila},Homb_inv,11,0),VLOOKUP(G{fila},Mujer_inv,11,0)))'
            hoja_calculo.cell(row=fila, column=27).value = f'=IF(D{fila}=\"F\",VLOOKUP(I{fila},HOMBRES,11,0),VLOOKUP(I{fila},MUJERES,11,0))'
            hoja_calculo.cell(row=fila, column=28).value = f'=IF(D{fila}=\"M\",VLOOKUP(G{fila},axy_,I{fila}-13,0),HLOOKUP(G{fila},axy_,I{fila}-13,0))'
            hoja_calculo.cell(row=fila, column=29).value = f'=AA{fila}-AB{fila}'
            hoja_calculo.cell(row=fila, column=30).value = f'=IF(U{fila}<5*$B$2,5*$B$2,IF(U{fila}>10*$B$2,10*$B$2,U{fila}))'
            hoja_calculo.cell(row=fila, column=31).value = f'=IF(D{fila}="M",VLOOKUP(G{fila},HOMBRES,14,0),VLOOKUP(G{fila},MUJERES,14,0))'
            hoja_calculo.cell(row=fila, column=32).value = f'=12*(Z{fila}*TABLA!$H$8+TABLA!$H$9)+2*(Z{fila}*TABLA!$H$10+TABLA!$H$11)'
            hoja_calculo.cell(row=fila, column=33).value = f'=12*(Z{fila}*TABLA!$H$8+TABLA!$H$9)+1*(Z{fila}*TABLA!$H$10+TABLA!$H$11)'
            hoja_calculo.cell(row=fila, column=34).value = f'=12*(Z{fila}*TABLA!$H$8+TABLA!$H$9)'
            hoja_calculo.cell(row=fila, column=35).value = f'=1*(Z{fila}*TABLA!$H$10+TABLA!$H$11)'
            hoja_calculo.cell(row=fila, column=36).value = f'=(12*AC{fila}*TABLA!$H$8)+1*AC{fila}*TABLA!$H$10'
            hoja_calculo.cell(row=fila, column=37).value = f'=(12*AC{fila}*TABLA!$H$8)+1*AC{fila}*TABLA!$H$10'
            hoja_calculo.cell(row=fila, column=38).value = f'=(12*AC{fila}*TABLA!$H$8)'
            hoja_calculo.cell(row=fila, column=39).value = f'=(1*AC{fila}*TABLA!$H$10)'
            hoja_calculo.cell(row=fila, column=40).value = f'=MAX(0,U{fila})*IF(V{fila}=14,AF{fila},AG{fila})'
            hoja_calculo.cell(row=fila, column=41).value = f'=AI{fila}*X{fila}'
            hoja_calculo.cell(row=fila, column=42).value = f'=IF(R{fila}=0,0,(U{fila}*IF(V{fila}=14,AJ{fila},AK{fila})))'
            hoja_calculo.cell(row=fila, column=43).value = f'=AM{fila}*X{fila}'
            hoja_calculo.cell(row=fila, column=44).value = f'=AH{fila}*Y{fila}'
            hoja_calculo.cell(row=fila, column=45).value = f'=AL{fila}*Y{fila}'
            hoja_calculo.cell(row=fila, column=46).value = f'=(AD{fila}*AE{fila})'
            hoja_calculo.cell(row=fila, column=47).value = f'=IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(P{fila},HOMBRES,11,0),VLOOKUP(P{fila},MUJERES,11,0)),IF(D{fila}="M",VLOOKUP(P{fila},Homb_inv,11,0),VLOOKUP(P{fila},Mujer_inv,11,0)))'
            hoja_calculo.cell(row=fila, column=48).value = f'=IF(E{fila}="NO",IF(D{fila}="M",VLOOKUP(P{fila},HOMBRES,11,0),VLOOKUP(P{fila},MUJERES,11,0)),IF(D{fila}="M",VLOOKUP(P{fila},Homb_inv,11,0),VLOOKUP(P{fila},Mujer_inv,11,0)))'
            hoja_calculo.cell(row=fila, column=49).value = f'=IF(D{fila}="M",VLOOKUP(P{fila},HOMBRES,3,0)/VLOOKUP(G{fila},HOMBRES,3,0),VLOOKUP(P{fila},MUJERES,3,0)/VLOOKUP(G{fila},MUJERES,3,0))'
            hoja_calculo.cell(row=fila, column=50).value = f'=IF(D{fila}="F",VLOOKUP(I{fila},HOMBRES,11,0),VLOOKUP(I{fila},MUJERES,11,0))'
            hoja_calculo.cell(row=fila, column=51).value = f'=IF(D{fila}="M",VLOOKUP(G{fila},axy_,I{fila}-13,0),HLOOKUP(G{fila},axy_,I{fila}-13,0))'
            hoja_calculo.cell(row=fila, column=52).value = f'=IF(D{fila}="M",VLOOKUP(P{fila},Dxy,I{fila}-13,0)/VLOOKUP(G{fila},Dxy,I{fila}-13,0),HLOOKUP(P{fila},Dxy,I{fila}-13,0)/HLOOKUP(G{fila},Dxy,I{fila}-13,0))'
            hoja_calculo.cell(row=fila, column=53).value = f'=AX{fila}-AY{fila}'
            hoja_calculo.cell(row=fila, column=54).value = f'=IF(S{fila}<5*$B$2,5*$B$2,IF(S{fila}>10*$B$2,10*$B$2,S{fila}))'
            hoja_calculo.cell(row=fila, column=55).value = f'=IF(D{fila}="M",VLOOKUP(G{fila},HOMBRES,14,0),VLOOKUP(G{fila},MUJERES,14,0))'
            hoja_calculo.cell(row=fila, column=56).value = f'=12*(AV{fila}*TABLA!$H$8+TABLA!$H$9)+2*(AV{fila}*TABLA!$H$10+TABLA!$H$11)'
            hoja_calculo.cell(row=fila, column=57).value = f'=12*(AV{fila}*TABLA!$H$8+TABLA!$H$9)+1*(AV{fila}*TABLA!$H$10+TABLA!$H$11)'
            hoja_calculo.cell(row=fila, column=58).value = f'=(12*BA{fila}*TABLA!$H$8)+2*BA{fila}*TABLA!$H$10'
            hoja_calculo.cell(row=fila, column=59).value = f'=(12*BA{fila}*TABLA!$H$8)+1*BA{fila}*TABLA!$H$10'
            hoja_calculo.cell(row=fila, column=60).value = f'=S{fila}*IF(T{fila}=14,BD{fila},BE{fila})*AW{fila}'
            hoja_calculo.cell(row=fila, column=61).value = f'=S{fila}*IF(T{fila}=14,BF{fila},BG{fila})'
            hoja_calculo.cell(row=fila, column=62).value = f'=(BB{fila}*BC{fila})'
            hoja_calculo.cell(row=fila, column=63).value = f'=SUM(BH{fila}:BJ{fila})'
            hoja_calculo.cell(row=fila, column=64).value = '=((1-(1+i)^-(\'Jub en expectativa(3)\'!K5/12))/((1+i)^(1/12)-1))*\'Jub en expectativa(3)\'!M5'
            #hoja_calculo.cell(row=fila, column=64).value = f'=((1-(1+i)^-(\'Jub en expectativa(3)\'!K5/12))/((1+i)^(1/12)-1))*\'Jub en expectativa(3)\'!M5'
            hoja_calculo.cell(row=fila, column=65).value = f'=AU{fila}-BK{fila}+BL{fila}'
            hoja_calculo.cell(row=fila, column=67).value = f'=AN{fila}+AO{fila}'
            hoja_calculo.cell(row=fila, column=68).value = f'=AP{fila}+AQ{fila}'
            hoja_calculo.cell(row=fila, column=69).value = f'=AT{fila}'
            hoja_calculo.cell(row=fila, column=70).value = f'=SUM(BO{fila}:BQ{fila})'
            hoja_calculo.cell(row=fila, column=71).value = f'=BH{fila}'
            hoja_calculo.cell(row=fila, column=72).value = f'=BI{fila}'
            hoja_calculo.cell(row=fila, column=73).value = f'=BJ{fila}'
            hoja_calculo.cell(row=fila, column=74).value = f'=BK{fila}'
            hoja_calculo.cell(row=fila, column=75).value = f'=BL{fila}'
            hoja_calculo.cell(row=fila, column=76).value = f'=+AR{fila}+AS{fila}'
        
        # Guardar el archivo Excel
        print("Guardando los datos")
        libro_excel.save(ruta_destino + nombre_archivo + '.xlsx')

        end_time = pd.Timestamp.now()
        tiempo_ejecucion = (end_time - start_time).total_seconds() / 60
        print(f"Tiempo de ejecución: {tiempo_ejecucion} minutos")
        print("Libro guardado con éxito")

    except Exception as e:
        print(f"Error al escribir en el archivo Excel: {e}")

