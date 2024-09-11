import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def rentas_escribir(datos_rentas, ruta_destino, nombre_archivo, ubicacion_formato):
    start_time = pd.Timestamp.now()

    # Verificar la ruta completa del archivo de plantilla
    print(f"Ruta del archivo de plantilla: {ubicacion_formato}")
    
    # Intentar cargar la plantilla de Excel
    try:
        libro_excel = openpyxl.load_workbook(ubicacion_formato)
        print("Archivo de plantilla cargado correctamente.")
    except Exception as e:
        print(f"Error al cargar el archivo de plantilla: {e}")
        return  # Termina la función si no se puede cargar el archivo

    # Leer los datos del archivo CSV
    try:
        print("Lectura del archivo CSV")
        datos_rentas_csv = pd.read_csv(datos_rentas, sep=';', encoding='latin1')
        print("Archivo CSV cargado correctamente.")
    except Exception as e:
        print(f"Error al leer el archivo CSV: {e}")
        return

    # Ahora escribimos en el archivo Excel usando openpyxl directamente
    try:
        # Seleccionar la hoja de cálculo
        hoja_calculo = libro_excel['Rentas Temporales']

        # Escribir las primeras 5 columnas
        print("Escribiendo los datos (primeras 5 columnas)")
        for idx, row in datos_rentas_csv.iloc[:, :5].iterrows():
            for col_idx, value in enumerate(row, start=1):
                hoja_calculo.cell(row=idx+11, column=col_idx, value=value)

        # Escribir las columnas 6, 7 y 8
        print("Escribiendo las columnas 6, 7 y 8")
        for idx, row in datos_rentas_csv.iloc[:, 5:8].iterrows():
            for col_idx, value in enumerate(row, start=7):  # Empezamos desde la columna 7
                hoja_calculo.cell(row=idx+11, column=col_idx, value=value)

        # Calcular las fórmulas
        num_filas = len(datos_rentas_csv) + 10
        for fila in range(11, num_filas + 1):
            hoja_calculo.cell(row=fila, column=6).value = f'=($B$6-D{fila})/365.25'
            hoja_calculo.cell(row=fila, column=10).value = f'=IF(H{fila}=14,($E$6+$E$7),IF(H{fila}=13,($E$6+$E$8),$E$6))'
            hoja_calculo.cell(row=fila, column=11).value = f'=(1-((1+TABLA!$B$10)^(-(G{fila}-F{fila}))))/TABLA!$B$10'
            hoja_calculo.cell(row=fila, column=12).value = f'=(1+TABLA!$B$10)/$E$3'
            hoja_calculo.cell(row=fila, column=13).value = f'=IF((I{fila}*J{fila}*K{fila}*L{fila})<0,0,(I{fila}*J{fila}*K{fila}*L{fila}))'

        # Guardar el archivo Excel con las modificaciones
        libro_excel.save(ruta_destino + nombre_archivo + '.xlsx')

        end_time = pd.Timestamp.now()
        tiempo_ejecucion = (end_time - start_time).total_seconds() / 60
        print(f"Tiempo de ejecución: {tiempo_ejecucion} minutos")
        print("Libro guardado con éxito")
        
    except Exception as e:
        print(f"Error al escribir en el archivo Excel: {e}")
